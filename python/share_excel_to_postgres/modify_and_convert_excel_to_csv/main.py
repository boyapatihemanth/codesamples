import pandas as pd
from datetime import *
from openpyxl import load_workbook

excel_file = 'data.xlsx'
sheet_name = 'Sheet1'
csv_file = 'data.csv'

def get_current_date():
    now = datetime.now()
    return now

def identify_unwanted_data_in_excel(current_date):
    """Identify data older than 7 days and add delete marker"""
    wb = load_workbook(excel_file)
    sheet = wb.active
    for i in range(sheet.max_row):
        excel_date = (sheet.cell(i+1,2)).value #NOTE: openpyxl treates 1st cell as (1,1) not (0,0)
        if type(excel_date) is str:
            continue #Escape 1st line of file which contains heading
        diff_from_current_date = current_date - excel_date

        if diff_from_current_date.days > 7:
            sheet.cell(i+1, 2).value = 'X' #Delete marker
    wb.save(excel_file)

def remove_unwanted_data_from_excel():
    """Identify row with delete marker and delete that row"""
    i = 1
    wb = load_workbook(excel_file)
    sheet = wb.active
    while i <= sheet.max_row:
        if sheet.cell(row=i, column = 2).value == 'X':
            sheet.delete_rows(i, 1)
        else:
            i += 1
    wb.save(excel_file)

def convert_excel_to_csv(in_excelfile, in_sheetname, out_csvfile):
    data_xls = pd.read_excel(in_excelfile, in_sheetname, index_col = 0)
    data_xls.to_csv(out_csvfile, encoding = 'utf-8')


if __name__ == '__main__':
    current_date = get_current_date()
    identify_unwanted_data_in_excel(current_date)
    remove_unwanted_data_from_excel()
    convert_excel_to_csv(excel_file, sheet_name, csv_file)